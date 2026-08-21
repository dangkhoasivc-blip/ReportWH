#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
export_hangguic1_v2.py
- Quét folder tìm TẤT CẢ file "Bao cao hang gui C1 theo kho (DD-MM-YYYY).xlsm"
- Mỗi file → data/bc01/hangguic1_YYYYMMDD.json
- Cập nhật data/bc01/manifest.json (danh sách ngày sẵn có)
- Chỉ re-export file nếu JSON chưa có hoặc Excel mới hơn JSON
"""
import json, shutil, tempfile, os, re
from pathlib import Path
from datetime import datetime
import openpyxl

EXCEL_DIR = Path("/sessions/cool-epic-wright/mnt/Gui C1")
DASHBOARD_DIR = Path("/sessions/cool-epic-wright/mnt/sabeco-dashboard")
OUT_DIR = DASHBOARD_DIR / "data" / "bc01"

def find_all_excels(folder):
    pattern = re.compile(r'Bao cao hang gui C1 theo kho \((\d{2}-\d{2}-\d{4})\)\.xlsm$', re.IGNORECASE)
    results = []
    for f in folder.glob("*.xlsm"):
        if f.name.startswith("~$"): continue
        m = pattern.match(f.name)
        if m:
            dd, mm, yyyy = m.group(1).split("-")
            date_key = f"{yyyy}{mm}{dd}"  # YYYYMMDD
            results.append((date_key, f))
    return sorted(results, key=lambda x: x[0])

def to_int(v):
    try: return int(float(v)) if v is not None and str(v).strip() not in ("","nan","None") else 0
    except: return 0

def to_float(v):
    try: return round(float(v),4) if v is not None and str(v).strip() not in ("","nan","None") else 0.0
    except: return 0.0

def fmt_date_str(date_key):
    return f"{date_key[6:8]}/{date_key[4:6]}/{date_key[:4]}"

def read_dashboard(wb):
    ws = wb["Dashboard"]
    rows = list(ws.iter_rows(values_only=True))
    header_row = None
    for i, row in enumerate(rows):
        if any(v is not None and str(v).strip() == "(Tcos)" for v in row):
            header_row = i; break
    if header_row is None: header_row = 2
    tcos_rows, overall = [], {}
    for row in rows[header_row+1:]:
        if row[1] is None: continue
        name = str(row[1]).strip()
        if not name or name.lower() == "nan": continue
        if name.upper().startswith("TỔNG") or name.upper().startswith("TONG"):
            overall = {"avgDays":to_int(row[2]),"total":to_int(row[3]),
                       "m4":to_int(row[4]),"m5":to_int(row[5]),"m6":to_int(row[6]),
                       "m7":to_int(row[7]),"m8":to_int(row[8]),
                       "hangBan":to_int(row[9]),"hangKm":to_int(row[10]) if len(row)>10 else 0}
            continue
        tcos_rows.append({"tcos":name,"avgDays":to_int(row[2]),"total":to_int(row[3]),
                          "m4":to_int(row[4]),"m5":to_int(row[5]),"m6":to_int(row[6]),
                          "m7":to_int(row[7]),"m8":to_int(row[8]),
                          "hangBan":to_int(row[9]),"hangKm":to_int(row[10]) if len(row)>10 else 0})
    return tcos_rows, overall

def read_detail(wb):
    ws = wb["BCTH hàng gửi C1 theo kho"]
    rows = list(ws.iter_rows(values_only=True))
    data_start = 6
    for i, row in enumerate(rows):
        vals = [v for v in row if v is not None]
        joined = " ".join(str(v).lower() for v in vals)
        if "volume" in joined or ("két thùng" in joined and "tổng số" in joined):
            data_start = i + 2; break
    detail = []
    for row in rows[data_start:]:
        if len(row) < 8: continue
        tcos_val = str(row[2]).strip() if row[2] else ""
        if not tcos_val or tcos_val.lower() in ("nan","tổng cộng","total","tcos","đvt"): continue
        vol = to_int(row[7])
        if vol == 0: continue
        detail.append({
            "no": to_int(row[1]), "tcos": tcos_val,
            "maKho": str(row[3]).strip() if row[3] else "",
            "tenKho": str(row[4]).strip() if row[4] else "",
            "standard": to_int(row[5]), "avgAge": to_int(row[6]),
            "volume": vol, "soHD": to_int(row[8]),
            "totalStockTcos": to_int(row[9]), "pctVsTcos": to_float(row[10]),
        })
    return detail

def read_npp(wb, today):
    """Đọc chi tiết từng hóa đơn/mã hàng theo NPP (Tên C1) từ sheet 'Data hang gui'."""
    ws = wb["Data hang gui"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows: return []
    header = rows[0]
    # Tìm vị trí cột theo tên (phòng khi thứ tự cột thay đổi)
    def find_col(name):
        for i, h in enumerate(header):
            if h and str(h).strip() == name:
                return i
        return None
    idx = {
        "tenKho": find_col("TÊN KHO"), "maKho": find_col("MÃ KHO"),
        "tenC1": find_col("TÊN C1"), "maC1": find_col("MÃ C1"),
        "soHD": find_col("SỐ HÓA ĐƠN"), "ngayHD": find_col("NGÀY RA HÓA ĐƠN"),
        "maHang": find_col("MÃ HÀNG"), "tenHang": find_col("TÊN HÀNG"),
        "slConLai": find_col("SỐ LƯỢNG CÒN LẠI"), "dvt": find_col("ĐƠN VỊ TÍNH"),
    }
    if any(v is None for v in idx.values()):
        return []
    npp_rows = []
    for row in rows[1:]:
        sl = row[idx["slConLai"]]
        try:
            sl = float(sl) if sl is not None else 0
        except:
            sl = 0
        if not sl or sl <= 0: continue
        ngay_hd = row[idx["ngayHD"]]
        age_days = None
        if ngay_hd:
            try:
                d = datetime.strptime(str(ngay_hd).strip(), "%d/%m/%Y")
                age_days = (today - d).days
            except:
                age_days = None
        npp_rows.append({
            "tenKho": str(row[idx["tenKho"]]).strip() if row[idx["tenKho"]] else "",
            "maKho": str(row[idx["maKho"]]).strip() if row[idx["maKho"]] else "",
            "tenC1": str(row[idx["tenC1"]]).strip() if row[idx["tenC1"]] else "",
            "maC1": str(row[idx["maC1"]]).strip() if row[idx["maC1"]] else "",
            "soHD": str(row[idx["soHD"]]).strip() if row[idx["soHD"]] else "",
            "ngayHD": str(ngay_hd).strip() if ngay_hd else "",
            "maHang": str(row[idx["maHang"]]).strip() if row[idx["maHang"]] else "",
            "tenHang": str(row[idx["tenHang"]]).strip() if row[idx["tenHang"]] else "",
            "dvt": str(row[idx["dvt"]]).strip() if row[idx["dvt"]] else "",
            "slConLai": to_int(sl),
            "ageDays": age_days if age_days is not None else 0,
        })
    return npp_rows

def export_one(date_key, src_path, out_path):
    tmp = tempfile.mktemp(suffix=".xlsm")
    shutil.copy2(src_path, tmp)
    try:
        wb = openpyxl.load_workbook(tmp, keep_vba=True, data_only=True, read_only=True)
        tcos_rows, overall = read_dashboard(wb)
        detail_rows = read_detail(wb)
        today = datetime.strptime(fmt_date_str(date_key), "%d/%m/%Y")
        npp_rows = read_npp(wb, today)
        wb.close()
    finally:
        os.unlink(tmp)

    if not overall:
        tot = sum(r["total"] for r in tcos_rows) or 1
        overall = {"total":tot,"avgDays":int(sum(r["avgDays"]*r["total"] for r in tcos_rows)/tot),
                   "m4":sum(r["m4"] for r in tcos_rows),"m5":sum(r["m5"] for r in tcos_rows),
                   "m6":sum(r["m6"] for r in tcos_rows),"m7":sum(r["m7"] for r in tcos_rows),
                   "m8":sum(r["m8"] for r in tcos_rows),"hangBan":sum(r["hangBan"] for r in tcos_rows),
                   "hangKm":sum(r["hangKm"] for r in tcos_rows)}
    tot = overall.get("total",1) or 1
    overall["pctBan"] = f"{overall['hangBan']/tot*100:.1f}%"
    overall["pctKm"]  = f"{overall['hangKm']/tot*100:.1f}%"

    payload = {
        "dateKey": date_key,
        "reportDate": fmt_date_str(date_key),
        "generatedAt": datetime.now().strftime("%d/%m/%Y %H:%M"),
        "overall": overall,
        "tcosRows": tcos_rows,
        "detailRows": detail_rows,
        "nppRows": npp_rows,
    }
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    return payload

def main():
    excels = find_all_excels(EXCEL_DIR)
    if not excels:
        print("Không tìm thấy file Excel nào."); return

    manifest_path = OUT_DIR / "manifest.json"
    manifest = []

    for date_key, src in excels:
        out_path = OUT_DIR / f"hangguic1_{date_key}.json"
        # Chỉ re-export nếu JSON chưa có hoặc Excel mới hơn
        if out_path.exists() and src.stat().st_mtime <= out_path.stat().st_mtime:
            print(f"  Skip {src.name} (đã có JSON)")
            # Đọc meta từ JSON cũ
            with open(out_path) as f:
                d = json.load(f)
            manifest.append({"dateKey": date_key, "reportDate": fmt_date_str(date_key),
                              "total": d["overall"]["total"], "file": f"data/bc01/hangguic1_{date_key}.json"})
        else:
            print(f"  Export {src.name}...")
            payload = export_one(date_key, src, out_path)
            manifest.append({"dateKey": date_key, "reportDate": fmt_date_str(date_key),
                              "total": payload["overall"]["total"],
                              "file": f"data/bc01/hangguic1_{date_key}.json"})
            print(f"    → {len(payload['tcosRows'])} tcos, {len(payload['detailRows'])} kho, total={payload['overall']['total']:,}")

    # Cập nhật manifest
    manifest.sort(key=lambda x: x["dateKey"])
    with open(manifest_path, "w", encoding="utf-8") as f:
        json.dump({"dates": manifest, "latest": manifest[-1]["dateKey"] if manifest else None}, f, ensure_ascii=False, indent=2)

    print(f"\n✓ manifest.json: {len(manifest)} ngày")
    print(f"  Ngày mới nhất: {manifest[-1]['reportDate']}")

if __name__ == "__main__":
    main()
