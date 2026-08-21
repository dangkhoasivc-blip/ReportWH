# -*- coding: utf-8 -*-
"""
export_to_json.py
Doc file Excel "Canh bao % shelf life.xlsm" (sheet Data + QD49)
-> sinh ra data.json cho web dashboard multi-report.

Rule canh bao lay dung theo file Excel (cot A sheet Data):
    =IFERROR(IF(B<=VLOOKUP(S,QD49!E:H,4,0),1,""),"")
    voi B = ROUNDDOWN(Q,0)  ->  floor(%HSD)
    S = SO NGAY HSD  ->  tra bang QD49 ra "% canh bao"
Nghia la: canh bao khi floor(%HSD) <= nguong cua chinh nhom han dung do
    182 ngay -> 32% | 273 ngay -> 27% | 365 ngay -> 20%

Chay:  python export_to_json.py
"""
import os
import re
import json
import math
import shutil
import tempfile
import datetime
import unicodedata

import openpyxl

# ============================================================
# CAU HINH DUONG DAN (tu dong nhan dien, khong can sua)
# ============================================================
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

EXCEL_CANDIDATES = [
    r"D:\OneDrive - SABECO\Cong viec\BC tuan\Canh bao % shelf life\Auto % Shelf life\Cảnh báo % shelf life.xlsm",
    r"D:\OneDrive - SABECO\Cong viec\BC tuan\Canh bao % shelf life\Cảnh báo % shelf life.xlsx",
    os.path.join(SCRIPT_DIR, "Cảnh báo % shelf life.xlsm"),
]

OUTPUT_JSON = os.path.join(SCRIPT_DIR, "data.json")

SHEET_DATA = "Data"
SHEET_RULE = "QĐ49"

# Chi so cot (0-based) tren sheet Data
C_DONVI, C_TENKHO, C_MAKHO = 2, 3, 4
C_MAHANG, C_TENHANG, C_DVT = 5, 6, 7
C_NHOM, C_TRANGTHAI, C_SOLO = 8, 9, 10
C_NSX, C_HSD = 12, 13
C_SOLUONG, C_SOLUONGPL = 14, 15
C_PCTHSD, C_NGAYCONLAI, C_NGAYHSD = 16, 17, 18
C_VITRI = 24

# Rut gon ten don vi dai -> ten khu vuc ngan (khop sheet "Vai tro")
REGION_MAP = {
    "CTY TNHH MTV TẬP ĐOÀN BIA SÀI GÒN": "Tập Đoàn BSG",
    "TRUNG TÂM PHÂN PHỐI CỦ CHI": "Củ Chi",
    "CTY CPTM BIA SÀI GÒN MIỀN BẮC": "Miền Bắc",
    "CTY CPTM BIA SÀI GÒN ĐÔNG BẮC": "Đông Bắc",
    "CTY CPTM BIA SÀI GÒN BẮC TRUNG BỘ": "Bắc Trung Bộ",
    "CTY CPTM BIA SÀI GÒN MIỀN TRUNG": "Miền Trung",
    "CTY CPTM BIA SÀI GÒN TÂY NGUYÊN": "Tây Nguyên",
    "CTY CPTM BIA SÀI GÒN NAM TRUNG BỘ": "Nam Trung Bộ",
    "CTY CPTM BIA SÀI GÒN MIỀN ĐÔNG": "Miền Đông",
    "CTY CPTM BIA SÀI GÒN TRUNG TÂM": "Trung Tâm",
    "CTY CPTM BIA SÀI GÒN SÔNG TIỀN": "Sông Tiền",
    "CTY CPTM BIA SÀI GÒN SÔNG HẬU": "Sông Hậu",
}

# Nguong mac dinh neu khong doc duoc sheet QD49
FALLBACK_RULE = {182: 32, 273: 27, 365: 20}


def log(msg):
    print(f"  {msg}", flush=True)


def find_excel():
    for p in EXCEL_CANDIDATES:
        if os.path.exists(p):
            return p
    raise FileNotFoundError(
        "Khong tim thay file Excel nguon. Kiem tra lai duong dan trong EXCEL_CANDIDATES."
    )


def norm_text(v):
    if v is None:
        return ""
    return str(v).strip()


def norm_key(v):
    """Bo dau + viet hoa de so sanh ten don vi."""
    nfkd = unicodedata.normalize("NFKD", norm_text(v).upper())
    return "".join(c for c in nfkd if not unicodedata.combining(c))


REGION_LOOKUP = {norm_key(k): v for k, v in REGION_MAP.items()}


def to_region(donvi):
    return REGION_LOOKUP.get(norm_key(donvi), norm_text(donvi))


def to_num(v):
    if isinstance(v, bool):
        return 0
    if isinstance(v, (int, float)):
        return v
    s = norm_text(v).replace(",", "")
    if not s:
        return 0
    try:
        return float(s)
    except ValueError:
        return 0


def fmt_date(v):
    """Cot NSX/HSD lan ca datetime va string dd/mm/yyyy -> chuan hoa ve dd/mm/yyyy."""
    if v is None or v == "":
        return ""
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%d/%m/%Y")
    s = norm_text(v)
    m = re.match(r"^(\d{1,2})[/-](\d{1,2})[/-](\d{4})", s)
    if m:
        d, mth, y = m.groups()
        return f"{int(d):02d}/{int(mth):02d}/{y}"
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})", s)
    if m:
        y, mth, d = m.groups()
        return f"{d}/{mth}/{y}"
    return s


def date_sort_key(s):
    """dd/mm/yyyy -> yyyymmdd de sort dung tren web."""
    m = re.match(r"^(\d{2})/(\d{2})/(\d{4})$", s or "")
    if not m:
        return 0
    d, mth, y = m.groups()
    return int(y) * 10000 + int(mth) * 100 + int(d)


def read_rule(wb):
    """Doc sheet QD49 -> {so_ngay_hsd: nguong_phan_tram}"""
    if SHEET_RULE not in wb.sheetnames:
        log(f"[CANH BAO] Khong thay sheet {SHEET_RULE}, dung nguong mac dinh.")
        return dict(FALLBACK_RULE)

    ws = wb[SHEET_RULE]
    rule = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 8:
            continue
        days, pct = row[4], row[7]   # E: So ngay HSD | H: % canh bao
        if isinstance(days, (int, float)) and isinstance(pct, (int, float)):
            rule[int(days)] = float(pct)
    if not rule:
        return dict(FALLBACK_RULE)
    return rule


def read_report_date(wb):
    """Sheet Dashboard o B2 chua 'Date report'."""
    if "Dashboard" in wb.sheetnames:
        ws = wb["Dashboard"]
        v = ws.cell(row=2, column=2).value
        if isinstance(v, (datetime.datetime, datetime.date)):
            return v.strftime("%d/%m/%Y")
        if v:
            return fmt_date(v)
    return datetime.datetime.now().strftime("%d/%m/%Y")


def run():
    print("=" * 62)
    print("  EXPORT DU LIEU SHELF LIFE -> data.json")
    print(f"  {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 62)

    excel_path = find_excel()
    log(f"File nguon: {excel_path}")

    # Copy ra temp de tranh loi file dang mo/bi lock
    tmp = os.path.join(tempfile.gettempdir(), "tmp_shelflife_export" + os.path.splitext(excel_path)[1])
    shutil.copy2(excel_path, tmp)

    try:
        wb = openpyxl.load_workbook(tmp, data_only=True, read_only=True)

        rule = read_rule(wb)
        log(f"Nguong canh bao QD49: " + ", ".join(f"{k} ngay <= {v}%" for k, v in sorted(rule.items())))

        report_date = read_report_date(wb)
        log(f"Ngay du lieu (Date report): {report_date}")

        ws = wb[SHEET_DATA]
        rows = []
        skipped = 0

        for raw in ws.iter_rows(min_row=2, values_only=True):
            if raw is None or len(raw) <= C_NGAYHSD:
                continue
            ma_hang = norm_text(raw[C_MAHANG])
            if not ma_hang:
                skipped += 1
                continue

            pct = to_num(raw[C_PCTHSD])
            days_hsd = int(to_num(raw[C_NGAYHSD]))
            pct_floor = math.floor(pct)

            threshold = rule.get(days_hsd)
            if threshold is None and rule:
                # Nhom han dung la, lay nguong cua nhom gan nhat
                nearest = min(rule.keys(), key=lambda k: abs(k - days_hsd))
                threshold = rule[nearest]
            threshold = threshold if threshold is not None else 0

            nsx = fmt_date(raw[C_NSX])
            hsd = fmt_date(raw[C_HSD])

            rows.append({
                "donVi": norm_text(raw[C_DONVI]),
                "khuVuc": to_region(raw[C_DONVI]),
                "tenKho": norm_text(raw[C_TENKHO]),
                "maKho": norm_text(raw[C_MAKHO]),
                "maHang": ma_hang,
                "tenHang": norm_text(raw[C_TENHANG]),
                "dvt": norm_text(raw[C_DVT]),
                "nhomHang": norm_text(raw[C_NHOM]),
                "trangThai": norm_text(raw[C_TRANGTHAI]),
                "soLo": norm_text(raw[C_SOLO]),
                "viTri": norm_text(raw[C_VITRI]) if len(raw) > C_VITRI else "",
                "nsx": nsx,
                "hsd": hsd,
                "nsxKey": date_sort_key(nsx),
                "hsdKey": date_sort_key(hsd),
                "soLuong": int(round(to_num(raw[C_SOLUONG]))),
                "soLuongPL": round(to_num(raw[C_SOLUONGPL]), 3),
                "pctHSD": round(pct, 2),
                "pctFloor": pct_floor,
                "ngayConLai": int(round(to_num(raw[C_NGAYCONLAI]))),
                "ngayHSD": days_hsd,
                "nguong": threshold,
                # Co canh bao theo dung rule QD49 hay khong
                "canhBao": 1 if pct_floor <= threshold else 0,
            })

        wb.close()
    finally:
        try:
            os.remove(tmp)
        except OSError:
            pass

    log(f"Doc duoc {len(rows)} dong du lieu (bo qua {skipped} dong trong).")

    canh_bao = [r for r in rows if r["canhBao"] == 1]
    log(f"So dong cham nguong QD49: {len(canh_bao)}")

    # Danh sach gia tri cho bo loc — sap xep tu nhien
    def uniq_sorted(key):
        return sorted({r[key] for r in rows if r[key]})

    filters = {
        "khuVuc": uniq_sorted("khuVuc"),
        "maKho": uniq_sorted("maKho"),
        "nhomHang": uniq_sorted("nhomHang"),
        "dvt": uniq_sorted("dvt"),
        "trangThai": uniq_sorted("trangThai"),
    }
    log("Bo loc: " + ", ".join(f"{k}={len(v)}" for k, v in filters.items()))

    now = datetime.datetime.now()
    payload = {
        "metadata": {
            "lastUpdatedDate": now.strftime("%d/%m/%Y"),
            "lastUpdatedTime": now.strftime("%H:%M:%S"),
            "reportDate": report_date,
            "sourceFile": os.path.basename(excel_path),
            "totalRows": len(rows),
            "rule": {str(k): v for k, v in sorted(rule.items())},
        },
        "filters": filters,
        "shelfLife": rows,
    }

    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, separators=(",", ":"))

    size_kb = os.path.getsize(OUTPUT_JSON) / 1024
    print()
    print("=" * 62)
    log(f"[OK] Da xuat: {OUTPUT_JSON}  ({size_kb:,.0f} KB)")
    print("=" * 62)


if __name__ == "__main__":
    try:
        run()
    except Exception as exc:
        print(f"\n  [LOI] {exc}")
        import traceback
        traceback.print_exc()
    print()
