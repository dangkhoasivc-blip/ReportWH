# -*- coding: utf-8 -*-
"""
export_hangblock.py
Doc file Excel "BC Hang Block.xlsm" (sheet Data + Dashboard)
-> sinh ra data_hangblock.json cho web dashboard multi-report.

Rule nghiep vu (lay tu fetch_api_hangblock.py):
    hang block = status IN ('D','R')
    D = Damaged (hu hong) | R = Rejected (bi tu choi)
Moi dong trong sheet Data DA LA hang block -> khong filter them.

QUAN TRONG: 2 cot nhap tay "Nguyen Nhan Block" va "Ke Hoach Clear"
CHI CO tren sheet Dashboard -> phai join Data <- Dashboard theo (MA HANG, SO LO).

Chay:  python export_hangblock.py
"""
import os
import re
import json
import shutil
import tempfile
import datetime
import unicodedata

import openpyxl

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

EXCEL_CANDIDATES = [
    r"D:\OneDrive - SABECO\Cong viec\BC tuan\Hang block\Nam 2026\Auto hàng block\BC Hàng Block.xlsm",
    os.path.join(SCRIPT_DIR, "BC Hàng Block.xlsm"),
]

OUTPUT_JSON = os.path.join(SCRIPT_DIR, "data_hangblock.json")

SHEET_DATA = "Data"
SHEET_DASH = "Dashboard"

# --- Chi so cot sheet Data (0-based) ---
D_NGAYTHAYDOI = 0
D_KEY = 1
D_NGAYTAO = 3
D_DONVI, D_TENKHO, D_MAKHO = 4, 5, 6
D_MAHANG, D_TENHANG, D_DVT = 7, 8, 9
D_NHOM, D_TRANGTHAI, D_SOLO = 10, 11, 12
D_NSX, D_HSD = 14, 15
D_SOLUONG, D_SOLUONGPL = 16, 17
D_PCTHSD, D_NGAYCONLAI, D_NGAYHSD = 18, 19, 20
D_VITRI = 26

# --- Chi so cot sheet Dashboard (0-based), header dong 5, data tu dong 6 ---
B_DONVI, B_TENKHO, B_NGAYTD = 1, 2, 3
B_MAHANG, B_SOLO, B_NSX = 4, 5, 6
B_TRANGTHAI, B_NHOM, B_SL = 7, 8, 9
B_NGUYENNHAN, B_KEHOACH = 10, 11

REGION_MAP = {
    "CTY TNHH MTV TẬP ĐOÀN BIA SÀI GÒN": "Tập Đoàn BSG",
    "TRUNG TÂM PHÂN PHỐI CỦ CHI": "Củ Chi",
    "CTY CPTM BIA SÀI GÒN MIỀN BẮC": "Miền Bắc",
    "CTY CPTM BIA SÀI GÒN ĐÔNG BẮC": "Đông Bắc",
    "CTY CPTM BIA SÀI GÒN BẮC TRUNG BỘ": "Bắc Trung Bộ",
    "CTY CPTM BIA SÀI GÒN MIỀN TRUNG": "Miền Trung",
    "CTY CPTM BIA SÀI GÒN TÂY NGUYÊN": "Tây Nguyên",
    "CTY CPTM BIA SÀI GÒN NAM TRUNG BỘ": "Nam Trung Bộ",
    "CTY CPTM BIA SÀI GÒN MIỀN ĐÔNG": "Miền Đông",
    "CTY CPTM BIA SÀI GÒN TRUNG TÂM": "Trung Tâm",
    "CTY CPTM BIA SÀI GÒN SÔNG TIỀN": "Sông Tiền",
    "CTY CPTM BIA SÀI GÒN SÔNG HẬU": "Sông Hậu",
}

TRANGTHAI_LABEL = {"D": "Hư hỏng (Damaged)", "R": "Bị từ chối (Rejected)"}
NHOM_LABEL = {"BB": "Bao bì", "TP": "Thành phẩm"}


def log(msg):
    print(f"  {msg}", flush=True)


def find_excel():
    for p in EXCEL_CANDIDATES:
        if os.path.exists(p):
            return p
    raise FileNotFoundError(
        "Khong tim thay file 'BC Hang Block.xlsm'. Kiem tra EXCEL_CANDIDATES."
    )


def norm_text(v):
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s in ("None", "nan") else s


def norm_key(v):
    nfkd = unicodedata.normalize("NFKD", norm_text(v).upper())
    return "".join(c for c in nfkd if not unicodedata.combining(c))


REGION_LOOKUP = {norm_key(k): v for k, v in REGION_MAP.items()}


def to_region(donvi):
    return REGION_LOOKUP.get(norm_key(donvi), norm_text(donvi))


def to_num(v):
    if isinstance(v, bool):
        return 0
    if isinstance(v, (int, float)):
        return v
    s = norm_text(v).replace(",", "")
    if not s:
        return 0
    try:
        return float(s)
    except ValueError:
        return 0


def fmt_date(v):
    """Cot ngay lan ca datetime va string dd/mm/yyyy -> chuan hoa dd/mm/yyyy."""
    if v is None or v == "":
        return ""
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%d/%m/%Y")
    s = norm_text(v)
    if not s:
        return ""
    m = re.match(r"^(\d{1,2})[/-](\d{1,2})[/-](\d{4})", s)
    if m:
        d, mth, y = m.groups()
        return f"{int(d):02d}/{int(mth):02d}/{y}"
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})", s)
    if m:
        y, mth, d = m.groups()
        return f"{d}/{mth}/{y}"
    return s


def date_sort_key(s):
    m = re.match(r"^(\d{2})/(\d{2})/(\d{4})$", s or "")
    if not m:
        return 0
    d, mth, y = m.groups()
    return int(y) * 10000 + int(mth) * 100 + int(d)


def fmt_plan(v):
    """Ke Hoach Clear: co the la datetime hoac text 'Pending'."""
    if v is None or v == "":
        return ""
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%d/%m/%Y")
    return norm_text(v)


def join_key(ma_hang, so_lo):
    """Key join Data <-> Dashboard."""
    return f"{norm_key(ma_hang)}||{norm_key(so_lo)}"


def read_dashboard(wb):
    """Doc sheet Dashboard -> {(mahang,solo): {nguyenNhan, keHoach}} + ngay bao cao."""
    manual = {}
    report_date = ""

    if SHEET_DASH not in wb.sheetnames:
        log(f"[CANH BAO] Khong thay sheet {SHEET_DASH}.")
        return manual, report_date

    ws = wb[SHEET_DASH]
    rows = list(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 500), values_only=True))

    # Data report o B3 (row 3, col index 2)
    if len(rows) >= 3 and len(rows[2]) > 2:
        report_date = fmt_date(rows[2][2])

    # Data tu dong 6 (index 5)
    for raw in rows[5:]:
        if not raw or len(raw) <= B_KEHOACH:
            continue
        donvi = norm_text(raw[B_DONVI])
        # Bo dong Grand Total va dong trong
        if not donvi or norm_key(donvi).startswith("GRAND TOTAL"):
            continue
        ma_hang = norm_text(raw[B_MAHANG])
        if not ma_hang:
            continue
        manual[join_key(ma_hang, raw[B_SOLO])] = {
            "nguyenNhan": norm_text(raw[B_NGUYENNHAN]),
            "keHoach": fmt_plan(raw[B_KEHOACH]),
        }

    return manual, report_date


def run():
    print("=" * 62)
    print("  EXPORT DU LIEU HANG BLOCK -> data_hangblock.json")
    print(f"  {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 62)

    excel_path = find_excel()
    log(f"File nguon: {excel_path}")

    tmp = os.path.join(tempfile.gettempdir(),
                       "tmp_hangblock_export" + os.path.splitext(excel_path)[1])
    shutil.copy2(excel_path, tmp)

    try:
        wb = openpyxl.load_workbook(tmp, data_only=True, read_only=True)

        manual, report_date = read_dashboard(wb)
        log(f"Sheet Dashboard: {len(manual)} dong co Nguyen Nhan / Ke Hoach Clear")
        if not report_date:
            report_date = datetime.datetime.now().strftime("%d/%m/%Y")
        log(f"Ngay du lieu (Data report): {report_date}")

        ws = wb[SHEET_DATA]
        rows = []
        skipped = 0
        matched = 0

        for raw in ws.iter_rows(min_row=2, values_only=True):
            if raw is None or len(raw) <= D_NGAYHSD:
                continue
            ma_hang = norm_text(raw[D_MAHANG])
            if not ma_hang:
                skipped += 1
                continue

            so_lo = norm_text(raw[D_SOLO])
            k = join_key(ma_hang, so_lo)
            info = manual.get(k, {})
            if info:
                matched += 1

            ngay_td = fmt_date(raw[D_NGAYTHAYDOI])
            nsx = fmt_date(raw[D_NSX])
            hsd = fmt_date(raw[D_HSD])
            tt = norm_text(raw[D_TRANGTHAI]).upper()
            nhom = norm_text(raw[D_NHOM]).upper()

            rows.append({
                "donVi": norm_text(raw[D_DONVI]),
                "khuVuc": to_region(raw[D_DONVI]),
                "tenKho": norm_text(raw[D_TENKHO]),
                "maKho": norm_text(raw[D_MAKHO]),
                "maHang": ma_hang,
                "tenHang": norm_text(raw[D_TENHANG]),
                "dvt": norm_text(raw[D_DVT]),
                "nhomHang": nhom,
                "nhomHangLabel": NHOM_LABEL.get(nhom, nhom),
                "trangThai": tt,
                "trangThaiLabel": TRANGTHAI_LABEL.get(tt, tt),
                "soLo": so_lo,
                "viTri": norm_text(raw[D_VITRI]) if len(raw) > D_VITRI else "",
                "ngayThayDoi": ngay_td,
                "ngayThayDoiKey": date_sort_key(ngay_td),
                "nsx": nsx,
                "hsd": hsd,
                "nsxKey": date_sort_key(nsx),
                "hsdKey": date_sort_key(hsd),
                "soLuong": int(round(to_num(raw[D_SOLUONG]))),
                "soLuongPL": round(to_num(raw[D_SOLUONGPL]), 4),
                "pctHSD": round(to_num(raw[D_PCTHSD]), 2),
                "ngayConLai": int(round(to_num(raw[D_NGAYCONLAI]))),
                "ngayHSD": int(round(to_num(raw[D_NGAYHSD]))),
                # 2 cot nhap tay tu sheet Dashboard
                "nguyenNhan": info.get("nguyenNhan", ""),
                "keHoach": info.get("keHoach", ""),
            })

        wb.close()
    finally:
        try:
            os.remove(tmp)
        except OSError:
            pass

    log(f"Doc duoc {len(rows)} dong hang block (bo qua {skipped} dong trong).")
    log(f"Join duoc Nguyen Nhan/Ke Hoach cho {matched}/{len(rows)} dong.")

    n_d = sum(1 for r in rows if r["trangThai"] == "D")
    n_r = sum(1 for r in rows if r["trangThai"] == "R")
    tong_sl = sum(r["soLuong"] for r in rows)
    log(f"Trang thai: D={n_d}, R={n_r} | Tong so luong: {tong_sl:,}")

    def uniq_sorted(key):
        return sorted({r[key] for r in rows if r[key]})

    filters = {
        "khuVuc": uniq_sorted("khuVuc"),
        "maKho": uniq_sorted("maKho"),
        "nhomHang": uniq_sorted("nhomHang"),
        "dvt": uniq_sorted("dvt"),
        "trangThai": uniq_sorted("trangThai"),
    }
    log("Bo loc: " + ", ".join(f"{k}={len(v)}" for k, v in filters.items()))

    now = datetime.datetime.now()
    payload = {
        "metadata": {
            "lastUpdatedDate": now.strftime("%d/%m/%Y"),
            "lastUpdatedTime": now.strftime("%H:%M:%S"),
            "reportDate": report_date,
            "sourceFile": os.path.basename(excel_path),
            "totalRows": len(rows),
            "tongSoLuong": tong_sl,
            "soDongD": n_d,
            "soDongR": n_r,
        },
        "filters": filters,
        "hangBlock": rows,
    }

    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, separators=(",", ":"))

    size_kb = os.path.getsize(OUTPUT_JSON) / 1024
    
    # Upload to Firebase
    try:
        from utils_firebase import upload_to_firebase
        print("\nUploading to Firebase...")
        upload_to_firebase('data_hangblock', payload)
    except ImportError:
        print("\n(Firebase utility not found, skipping upload)")

    print(f"\n[OK] Xuất thành công: {OUTPUT_JSON} ({size_kb:.1f} KB)")
    print("=" * 62)


if __name__ == "__main__":
    try:
        run()
    except Exception as exc:
        print(f"\n  [LOI] {exc}")
        import traceback
        traceback.print_exc()
    print()
