# -*- coding: utf-8 -*-
"""
export_hopdong.py
Doc file Excel "Bao cao theo doi thoi han hop dong v2.2026..xlsx" (sheet Dashboard)
-> sinh ra data_hopdong.json cho web dashboard multi-report.

Sheet Dashboard: dong 1 = "Date report" (C1 = TODAY()), dong 2 = header, du lieu tu dong 3.
Cot TRANG THAI va THOI HAN HD CON LAI da duoc Excel tinh san bang cong thuc dua vao TODAY(),
nen script nay CHI DOC GIA TRI (data_only=True), khong tu tinh lai.

Gia tri TRANG THAI: On-Going | NearExpiry | Expired | "-" (hop dong khong xac dinh thoi han).

Join them sheet "Vai tro" (Khu vuc -> Truong Kho, Nhac Pic, Email) de hien thi nguoi phu trach.

Chay:  python export_hopdong.py
"""
import os
import re
import json
import shutil
import tempfile
import datetime

import openpyxl

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

EXCEL_CANDIDATES = [
    r"D:\OneDrive - SABECO\Portal Satraco - WH Report\Báo cáo theo dõi thời hạn hợp đồng\Báo cáo theo dõi thời hạn hợp đồng v2.2026..xlsx",
    os.path.join(SCRIPT_DIR, "Báo cáo theo dõi thời hạn hợp đồng v2.2026..xlsx"),
]

OUTPUT_JSON = os.path.join(SCRIPT_DIR, "data_hopdong.json")

SHEET_DASH = "Dashboard"
SHEET_ROLE = "Vai trò"

# Chi so cot sheet Dashboard (0-based), header o dong 2 (index 1), data tu dong 3 (index 2)
H_STT, H_KHUVUC, H_KHO = 0, 1, 2
H_HOPDONG, H_NCC = 3, 4
H_NGAYBD_CU, H_NGAYGH_CU, H_NGAYHH_CU = 5, 6, 7
H_NGAYBD_MOI, H_NGAYGH_MOI, H_NGAYHH_MOI = 8, 9, 10
H_THOIHAN, H_TRANGTHAI = 11, 12
H_NGAYTOTRINH, H_GHICHU = 13, 14

TRANGTHAI_LABEL = {
    "On-Going": "Còn hiệu lực",
    "NearExpiry": "Gần hết hạn (≤60 ngày)",
    "Expired": "Đã hết hạn",
    "-": "Không xác định thời hạn",
}


def log(msg):
    print(f"  {msg}", flush=True)


def find_excel():
    for p in EXCEL_CANDIDATES:
        if os.path.exists(p):
            return p
    raise FileNotFoundError(
        "Khong tim thay file 'Bao cao theo doi thoi han hop dong v2.2026..xlsx'. "
        "Kiem tra EXCEL_CANDIDATES."
    )


def norm_text(v):
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s in ("None", "nan", "-") else s


def norm_text_dash(v):
    """Giu nguyen dau '-' (dung cho Trang Thai / Thoi Han)."""
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s in ("None", "nan") else s


def fmt_date(v):
    """Cot ngay lan ca datetime va string dd/mm/yyyy -> chuan hoa dd/mm/yyyy."""
    if v is None or v == "":
        return ""
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%d/%m/%Y")
    s = norm_text(v)
    if not s:
        return ""
    m = re.match(r"^(\d{1,2})[/-](\d{1,2})[/-](\d{4})", s)
    if m:
        d, mth, y = m.groups()
        return f"{int(d):02d}/{int(mth):02d}/{y}"
    return s


def date_sort_key(s):
    m = re.match(r"^(\d{2})/(\d{2})/(\d{4})$", s or "")
    if not m:
        return 0
    d, mth, y = m.groups()
    return int(y) * 10000 + int(mth) * 100 + int(d)


def parse_so_ngay(thoi_han, trang_thai):
    """'136 NGÀY CÒN HIỆU LỰC' -> 136 | '139 NGÀY ĐÃ HẾT HẠN' -> -139 | '-' -> None."""
    m = re.match(r"^(\d+)\s*NGÀY", thoi_han or "", re.IGNORECASE)
    if not m:
        return None
    n = int(m.group(1))
    return -n if trang_thai == "Expired" else n


def read_roles(wb):
    """Doc sheet 'Vai tro' -> {khuVuc: {vaiTro, nhacPic, email}}."""
    roles = {}
    if SHEET_ROLE not in wb.sheetnames:
        return roles
    ws = wb[SHEET_ROLE]
    rows = list(ws.iter_rows(values_only=True))
    for raw in rows[1:]:
        if not raw or not raw[0]:
            continue
        khu_vuc = norm_text(raw[0])
        if not khu_vuc:
            continue
        roles[khu_vuc] = {
            "vaiTro": norm_text(raw[1]) if len(raw) > 1 else "",
            "nhacPic": norm_text(raw[2]) if len(raw) > 2 else "",
            "email": norm_text(raw[3]) if len(raw) > 3 else "",
        }
    return roles


def run():
    print("=" * 62)
    print("  EXPORT BC04 - THEO DOI THOI HAN HOP DONG -> data_hopdong.json")
    print(f"  {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 62)

    excel_path = find_excel()
    log(f"File nguon: {excel_path}")

    tmp = os.path.join(tempfile.gettempdir(),
                        "tmp_hopdong_export" + os.path.splitext(excel_path)[1])
    shutil.copy2(excel_path, tmp)

    try:
        wb = openpyxl.load_workbook(tmp, data_only=True, read_only=True)

        roles = read_roles(wb)
        log(f"Sheet '{SHEET_ROLE}': {len(roles)} khu vực")

        ws = wb[SHEET_DASH]
        rows = list(ws.iter_rows(values_only=True))

        report_date = ""
        if rows and len(rows[0]) > 2 and rows[0][2]:
            report_date = fmt_date(rows[0][2])
        if not report_date:
            report_date = datetime.datetime.now().strftime("%d/%m/%Y")
        log(f"Ngày báo cáo (Date report): {report_date}")

        data = []
        for raw in rows[2:]:
            if not raw or len(raw) <= H_TRANGTHAI:
                continue
            khu_vuc = norm_text(raw[H_KHUVUC])
            hop_dong = norm_text(raw[H_HOPDONG])
            if not khu_vuc and not hop_dong:
                continue

            trang_thai = norm_text_dash(raw[H_TRANGTHAI])
            thoi_han = norm_text_dash(raw[H_THOIHAN])
            role_info = roles.get(khu_vuc, {})

            data.append({
                "stt": raw[H_STT] if isinstance(raw[H_STT], (int, float)) else len(data) + 1,
                "khuVuc": khu_vuc,
                "kho": norm_text(raw[H_KHO]),
                "hopDong": hop_dong,
                "tenNCC": norm_text(raw[H_NCC]),
                "ngayBdCu": fmt_date(raw[H_NGAYBD_CU]),
                "ngayGhCu": fmt_date(raw[H_NGAYGH_CU]),
                "ngayHhCu": fmt_date(raw[H_NGAYHH_CU]),
                "ngayBdMoi": fmt_date(raw[H_NGAYBD_MOI]),
                "ngayGhMoi": fmt_date(raw[H_NGAYGH_MOI]),
                "ngayHhMoi": fmt_date(raw[H_NGAYHH_MOI]),
                "thoiHan": thoi_han,
                "trangThai": trang_thai,
                "trangThaiLabel": TRANGTHAI_LABEL.get(trang_thai, trang_thai),
                "soNgay": parse_so_ngay(thoi_han, trang_thai),
                "ngayToTrinh": fmt_date(raw[H_NGAYTOTRINH]),
                "ghiChu": norm_text(raw[H_GHICHU]),
                "vaiTro": role_info.get("vaiTro", ""),
                "nhacPic": role_info.get("nhacPic", ""),
                "email": role_info.get("email", ""),
            })

        wb.close()
    finally:
        try:
            os.remove(tmp)
        except OSError:
            pass

    log(f"Đọc được {len(data)} dòng hợp đồng.")

    n_ongoing = sum(1 for r in data if r["trangThai"] == "On-Going")
    n_near = sum(1 for r in data if r["trangThai"] == "NearExpiry")
    n_expired = sum(1 for r in data if r["trangThai"] == "Expired")
    n_undef = sum(1 for r in data if r["trangThai"] == "-")
    log(f"Trạng thái: On-Going={n_ongoing}, NearExpiry={n_near}, "
        f"Expired={n_expired}, Không xác định={n_undef}")

    def uniq_sorted(key):
        return sorted({r[key] for r in data if r[key]})

    filters = {
        "khuVuc": uniq_sorted("khuVuc"),
        "kho": uniq_sorted("kho"),
        "trangThai": uniq_sorted("trangThai"),
        "vaiTro": uniq_sorted("vaiTro"),
    }
    log("Bộ lọc: " + ", ".join(f"{k}={len(v)}" for k, v in filters.items()))

    now = datetime.datetime.now()
    payload = {
        "metadata": {
            "lastUpdatedDate": now.strftime("%d/%m/%Y"),
            "lastUpdatedTime": now.strftime("%H:%M:%S"),
            "reportDate": report_date,
            "sourceFile": os.path.basename(excel_path),
            "totalRows": len(data),
            "soOnGoing": n_ongoing,
            "soNearExpiry": n_near,
            "soExpired": n_expired,
        },
        "filters": filters,
        "hopDong": data,
    }

    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, separators=(",", ":"))

    size_kb = os.path.getsize(OUTPUT_JSON) / 1024
    print()
    print("=" * 62)
    log(f"[OK] Đã xuất: {OUTPUT_JSON}  ({size_kb:,.1f} KB)")
    print("=" * 62)


if __name__ == "__main__":
    try:
        run()
    except Exception as exc:
        print(f"\n  [LỖI] {exc}")
        import traceback
        traceback.print_exc()
    print()
